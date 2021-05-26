import openpyxl
from openpyxl.styles.fonts import Font
from make_plt import make_plt 

def output_graph(flist,list_data,userid):
    num = len(list_data)

    #グラフの出力場所（セル番号）
    cell_num=[9,24,39,5,20,35,50]

    wb = openpyxl.load_workbook('output_Templete.xlsx')
    sheet = wb.worksheets[0]
    ws = wb.active

    graph_array = [[0 for j in range(num)] for j in range(7)]
    for list_two in range(7):
        for list_one in range(num):
            graph_array[list_two][list_one] = (list_data[list_one][list_two])#配列のxとyを逆転
        
        img = make_plt(flist,graph_array[list_two],list_two,userid)
        img.width = 625
        img.height = 250
        if list_two < 3:
            row_number = cell_num[list_two]
            cell = ws.cell(row=row_number, column=2).coordinate#セルの位置を指定（column=3はC列のこと
            img.anchor = cell
            ws.add_image(img)
        else:
            row_number = cell_num[list_two]
            cell = ws.cell(row=row_number, column=14).coordinate
            img.anchor = cell
            ws.add_image(img)

    sheet['B4'].font = Font(size=14)
    sheet['B4'] = '%sさん の学習達成度推移です。'%userid
    wb.save('output/%s.xlsx'%userid)# output.pyで出力した.pngをout.xlsxに貼り付ける
